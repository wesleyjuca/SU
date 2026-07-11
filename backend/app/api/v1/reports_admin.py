"""Relatórios consolidados da banca (visão matriz das unidades do mesmo escritório).

Agrega processos, financeiro, IA e usuários de TODAS as unidades de uma banca
(tenant-mãe + tenants filhos via parent_tenant_id). Isolamento normal por
tenant_id é preservado — só quem tem direito à matriz enxerga o consolidado:
- SUPERADMIN: qualquer banca (via ?banca_id=).
- ADMIN/SÓCIO: apenas a própria banca, se ela tiver unidades.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from datetime import datetime, timezone
import csv
import io
import uuid

from app.db.base import get_db
from app.dependencies import require_role
from app.models.user import User
from app.models.tenant import Tenant
from app.models.process import LegalProcess
from app.models.financial import FinancialEntry
from app.models.agent_run import AgentRun
from app.utils.periods import inicio_mes, range_meses

router = APIRouter(prefix="/reports", tags=["reports-admin"])


async def _children(db: AsyncSession, banca_id: uuid.UUID) -> list[Tenant]:
    return list((await db.execute(
        select(Tenant).where(Tenant.parent_tenant_id == banca_id).order_by(Tenant.created_at)
    )).scalars().all())


async def _has_children(db: AsyncSession, tenant_id: uuid.UUID) -> bool:
    return (await db.execute(
        select(func.count(Tenant.id)).where(Tenant.parent_tenant_id == tenant_id)
    )).scalar_one() > 0


async def _resolve_banca(db: AsyncSession, current_user: User, banca_id: str | None):
    """Resolve (banca, unidades) respeitando o gate. Levanta 403/404 quando indevido."""
    if current_user.role == "SUPERADMIN":
        if not banca_id:
            raise HTTPException(status_code=422, detail="Informe a banca (banca_id).")
        banca = (await db.execute(
            select(Tenant).where(Tenant.id == uuid.UUID(banca_id))
        )).scalar_one_or_none()
        if not banca:
            raise HTTPException(status_code=404, detail="Banca não encontrada.")
        if banca.parent_tenant_id is not None:
            raise HTTPException(status_code=422, detail="Selecione uma banca-mãe, não uma unidade.")
    else:
        # ADMIN/SÓCIO: sempre a própria banca (ignora banca_id recebido).
        if not current_user.tenant_id:
            raise HTTPException(status_code=403, detail="Usuário sem escritório vinculado.")
        banca = (await db.execute(
            select(Tenant).where(Tenant.id == current_user.tenant_id)
        )).scalar_one_or_none()
        if not banca or banca.parent_tenant_id is not None:
            raise HTTPException(status_code=403, detail="Relatório disponível apenas para a matriz da banca.")

    unidades = await _children(db, banca.id)
    return banca, unidades


@router.get("/scope")
async def reports_scope(
    current_user: User = Depends(require_role("ADMIN", "SOCIO")),
    db: AsyncSession = Depends(get_db),
):
    """Descoberta para o frontend: se o usuário pode ver o consolidado e de quais bancas."""
    is_super = current_user.role == "SUPERADMIN"
    if is_super:
        # Todas as bancas-mãe (tenants com ao menos um filho).
        parents = (await db.execute(
            select(Tenant.parent_tenant_id).where(Tenant.parent_tenant_id.isnot(None)).distinct()
        )).scalars().all()
        bancas = []
        if parents:
            rows = (await db.execute(
                select(Tenant).where(Tenant.id.in_(parents)).order_by(Tenant.name)
            )).scalars().all()
            counts = dict((await db.execute(
                select(Tenant.parent_tenant_id, func.count(Tenant.id))
                .where(Tenant.parent_tenant_id.in_(parents)).group_by(Tenant.parent_tenant_id)
            )).all())
            bancas = [{"id": str(t.id), "name": t.name, "unidades": int(counts.get(t.id, 0))} for t in rows]
        return {"can_view": True, "is_superadmin": True, "bancas": bancas}

    # ADMIN/SÓCIO: só se o próprio tenant for banca-mãe.
    if not current_user.tenant_id:
        return {"can_view": False, "is_superadmin": False, "bancas": []}
    banca = (await db.execute(
        select(Tenant).where(Tenant.id == current_user.tenant_id)
    )).scalar_one_or_none()
    if not banca or banca.parent_tenant_id is not None or not await _has_children(db, banca.id):
        return {"can_view": False, "is_superadmin": False, "bancas": []}
    n = (await db.execute(
        select(func.count(Tenant.id)).where(Tenant.parent_tenant_id == banca.id)
    )).scalar_one()
    return {"can_view": True, "is_superadmin": False,
            "bancas": [{"id": str(banca.id), "name": banca.name, "unidades": int(n)}]}


@router.get("/consolidated")
async def consolidated_report(
    banca_id: str | None = Query(default=None),
    current_user: User = Depends(require_role("ADMIN", "SOCIO")),
    db: AsyncSession = Depends(get_db),
):
    """Matriz consolidada da banca — uma linha por unidade + totais + série mensal."""
    banca, unidades = await _resolve_banca(db, current_user, banca_id)
    return await _compute_consolidated(db, banca, unidades)


async def _compute_consolidated(db: AsyncSession, banca: Tenant, unidades: list[Tenant]) -> dict:
    """Computa a matriz consolidada (reusada pelo JSON e pela exportação)."""
    tenants = [banca, *unidades]
    tenant_ids = [t.id for t in tenants]
    nomes = {banca.id: f"{banca.name} (matriz)"}
    for u in unidades:
        nomes[u.id] = u.unit_label or u.name

    ini_mes = inicio_mes()

    # ── Processos por tenant (ativos e total) ────────────────────────────────
    proc_ativos = dict((await db.execute(
        select(LegalProcess.tenant_id, func.count(LegalProcess.id))
        .where(LegalProcess.tenant_id.in_(tenant_ids), LegalProcess.situacao == "ATIVO")
        .group_by(LegalProcess.tenant_id)
    )).all())
    proc_total = dict((await db.execute(
        select(LegalProcess.tenant_id, func.count(LegalProcess.id))
        .where(LegalProcess.tenant_id.in_(tenant_ids))
        .group_by(LegalProcess.tenant_id)
    )).all())

    # ── Financeiro do mês (receita/despesa pagas) por tenant × tipo ──────────
    # Filtra pela criação do lançamento (timestamptz), consistente com IA/uso.
    fin_rows = (await db.execute(
        select(FinancialEntry.tenant_id, FinancialEntry.tipo, func.coalesce(func.sum(FinancialEntry.valor), 0))
        .where(
            FinancialEntry.tenant_id.in_(tenant_ids),
            FinancialEntry.status == "PAGO",
            FinancialEntry.created_at >= ini_mes,
        )
        .group_by(FinancialEntry.tenant_id, FinancialEntry.tipo)
    )).all()
    receita: dict = {}
    despesa: dict = {}
    for tid, tipo, val in fin_rows:
        (receita if tipo == "RECEITA" else despesa)[tid] = float(val or 0)

    # ── IA do mês (custo + execuções) por tenant ─────────────────────────────
    ia_rows = (await db.execute(
        select(
            AgentRun.tenant_id,
            func.coalesce(func.sum(AgentRun.cost_usd), 0),
            func.count(AgentRun.id),
        )
        .where(AgentRun.tenant_id.in_(tenant_ids), AgentRun.started_at >= ini_mes)
        .group_by(AgentRun.tenant_id)
    )).all()
    ia_custo = {tid: float(c or 0) for tid, c, _ in ia_rows}
    ia_exec = {tid: int(n or 0) for tid, _, n in ia_rows}

    # ── Usuários ativos por tenant ───────────────────────────────────────────
    users_ativos = dict((await db.execute(
        select(User.tenant_id, func.count(User.id))
        .where(User.tenant_id.in_(tenant_ids), User.is_active == True)  # noqa: E712
        .group_by(User.tenant_id)
    )).all())

    linhas = []
    for t in tenants:
        rec = receita.get(t.id, 0.0)
        desp = despesa.get(t.id, 0.0)
        linhas.append({
            "tenant_id": str(t.id),
            "unidade": nomes[t.id],
            "is_matriz": t.id == banca.id,
            "processos_ativos": int(proc_ativos.get(t.id, 0)),
            "processos_total": int(proc_total.get(t.id, 0)),
            "receita_mes": rec,
            "despesa_mes": desp,
            "saldo_mes": rec - desp,
            "custo_ia_mes": ia_custo.get(t.id, 0.0),
            "execucoes_mes": ia_exec.get(t.id, 0),
            "usuarios_ativos": int(users_ativos.get(t.id, 0)),
        })

    totais = {
        "processos_ativos": sum(l["processos_ativos"] for l in linhas),
        "processos_total": sum(l["processos_total"] for l in linhas),
        "receita_mes": sum(l["receita_mes"] for l in linhas),
        "despesa_mes": sum(l["despesa_mes"] for l in linhas),
        "saldo_mes": sum(l["saldo_mes"] for l in linhas),
        "custo_ia_mes": sum(l["custo_ia_mes"] for l in linhas),
        "execucoes_mes": sum(l["execucoes_mes"] for l in linhas),
        "usuarios_ativos": sum(l["usuarios_ativos"] for l in linhas),
    }

    # ── Série mensal consolidada (receita×despesa, últimos 6 meses) ──────────
    desde = range_meses(6)
    mes = func.date_trunc("month", FinancialEntry.created_at)
    serie_rows = (await db.execute(
        select(mes.label("mes"), FinancialEntry.tipo, func.coalesce(func.sum(FinancialEntry.valor), 0))
        .where(
            FinancialEntry.tenant_id.in_(tenant_ids),
            FinancialEntry.status == "PAGO",
            FinancialEntry.created_at >= desde,
        )
        .group_by("mes", FinancialEntry.tipo).order_by("mes")
    )).all()
    serie_map: dict = {}
    for m, tipo, val in serie_rows:
        key = m.date().isoformat() if hasattr(m, "date") else str(m)
        entry = serie_map.setdefault(key, {"mes": key, "receita": 0.0, "despesa": 0.0})
        entry["receita" if tipo == "RECEITA" else "despesa"] = float(val or 0)
    series = sorted(serie_map.values(), key=lambda e: e["mes"])

    return {
        "banca": {"id": str(banca.id), "name": banca.name, "unidades": len(unidades)},
        "linhas": linhas,
        "totais": totais,
        "series": series,
        "gerado_em": datetime.now(timezone.utc).isoformat(),
    }


def _brl(v) -> str:
    return "R$ " + f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


async def _banca_letterhead(db: AsyncSession, banca: Tenant) -> dict | None:
    """Timbrado da banca-mãe (mesmo padrão de documents.py, com banca.id)."""
    from app.models.tenant import TenantConfig
    cfg = (await db.execute(
        select(TenantConfig).where(TenantConfig.tenant_id == banca.id)
    )).scalar_one_or_none()
    if not cfg:
        return None
    lh = dict((cfg.document_templates or {}).get("letterhead", {}))
    if cfg.logo_url:
        lh["logo_data_url"] = cfg.logo_url
    return lh


@router.get("/consolidated/export")
async def export_consolidated(
    format: str = Query(...),
    banca_id: str | None = Query(default=None),
    current_user: User = Depends(require_role("ADMIN", "SOCIO")),
    db: AsyncSession = Depends(get_db),
):
    """Exporta o consolidado da banca em PDF (timbrado), XLSX ou CSV."""
    fmt = (format or "").lower()
    if fmt not in ("pdf", "csv", "xlsx"):
        raise HTTPException(status_code=422, detail="Formato inválido. Use 'pdf', 'xlsx' ou 'csv'.")

    banca, unidades = await _resolve_banca(db, current_user, banca_id)
    report = await _compute_consolidated(db, banca, unidades)
    linhas, totais = report["linhas"], report["totais"]
    competencia = inicio_mes().strftime("%Y-%m")
    periodo_label = "Competência: " + inicio_mes().strftime("%m/%Y")
    fname = f"consolidado-{banca.slug}-{competencia}"

    if fmt == "csv":
        buf = io.StringIO()
        buf.write("﻿")  # BOM p/ o Excel pt-BR reconhecer UTF-8
        w = csv.writer(buf, delimiter=";")
        w.writerow(["Unidade", "Proc. ativos", "Proc. total", "Receita mes",
                    "Despesa mes", "Saldo mes", "Custo IA (USD)", "Execucoes", "Usuarios"])
        for l in linhas:
            w.writerow([l["unidade"], l["processos_ativos"], l["processos_total"],
                        _brl(l["receita_mes"]), _brl(l["despesa_mes"]), _brl(l["saldo_mes"]),
                        f"{l['custo_ia_mes']:.2f}", l["execucoes_mes"], l["usuarios_ativos"]])
        w.writerow(["TOTAL DA BANCA", totais["processos_ativos"], totais["processos_total"],
                    _brl(totais["receita_mes"]), _brl(totais["despesa_mes"]), _brl(totais["saldo_mes"]),
                    f"{totais['custo_ia_mes']:.2f}", totais["execucoes_mes"], totais["usuarios_ativos"]])
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'},
        )

    if fmt == "xlsx":
        xlsx = _build_xlsx(banca.name, periodo_label, linhas, totais)
        return Response(
            content=xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'},
        )

    # PDF
    from app.utils.pdf_builder import build_consolidated_pdf
    letterhead = await _banca_letterhead(db, banca)
    pdf = build_consolidated_pdf(banca.name, periodo_label, linhas, totais, letterhead)
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'},
    )


def _build_xlsx(banca_name: str, periodo_label: str, linhas: list[dict], totais: dict) -> bytes:
    """Planilha .xlsx real (openpyxl) da matriz consolidada, com estilos e total em negrito."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    navy = PatternFill("solid", fgColor="1E2229")
    gold = PatternFill("solid", fgColor="C9A84C")
    white_bold = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D2C4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    ws["A1"] = f"Relatório Consolidado — {banca_name}"
    ws["A1"].font = Font(bold=True, size=13, color="1E2229")
    ws["A2"] = periodo_label
    ws["A2"].font = Font(size=10, color="6B6B6B")

    header = ["Unidade", "Proc. ativos", "Proc. total", "Receita mês", "Despesa mês",
              "Saldo mês", "Custo IA (USD)", "Execuções", "Usuários"]
    hrow = 4
    for c, titulo in enumerate(header, start=1):
        cell = ws.cell(row=hrow, column=c, value=titulo)
        cell.fill = navy
        cell.font = white_bold
        cell.border = border
        cell.alignment = right if c > 1 else Alignment(horizontal="left")

    def _num(v) -> float:
        return round(float(v), 2)

    r = hrow + 1
    for l in linhas:
        vals = [l["unidade"], l["processos_ativos"], l["processos_total"],
                _num(l["receita_mes"]), _num(l["despesa_mes"]), _num(l["saldo_mes"]),
                _num(l["custo_ia_mes"]), l["execucoes_mes"], l["usuarios_ativos"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if c in (4, 5, 6):
                cell.number_format = 'R$ #,##0.00'
            elif c == 7:
                cell.number_format = '"$" #,##0.00'
        r += 1

    total_vals = ["TOTAL DA BANCA", totais["processos_ativos"], totais["processos_total"],
                  _num(totais["receita_mes"]), _num(totais["despesa_mes"]), _num(totais["saldo_mes"]),
                  _num(totais["custo_ia_mes"]), totais["execucoes_mes"], totais["usuarios_ativos"]]
    for c, v in enumerate(total_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = gold
        cell.font = white_bold
        cell.border = border
        if c in (4, 5, 6):
            cell.number_format = 'R$ #,##0.00'
        elif c == 7:
            cell.number_format = '"$" #,##0.00'

    widths = [34, 12, 12, 15, 15, 15, 14, 12, 10]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
